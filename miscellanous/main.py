import cv2
from plyer import notification
import datetime
import requests
import flask
from pyngrok import ngrok
import openpyxl
import logging

current_line = 2
mode = input('Type "H" to host a class, Type "J" to join a class > ')
if mode == 'H' or mode == 'h':
    app = flask.Flask(__name__)
    log = logging.getLogger('werkzeug')
    log.setLevel(logging.ERROR)
    participants = []
    topic = input('Ok...Enter class topic: ')

    while 1:
        try:
            path = input('Enter path to save attendance file: ')
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet['A1'] = 'Participants'
            sheet['B1'] = 'Roll Number'
            sheet['C1'] = 'Time Joined'
            sheet['D1'] = 'Time Left'
            workbook.save(f'{path}/{topic}_attendance.xlsx')
            main_book = openpyxl.load_workbook(f'{path}/{topic}_attendance.xlsx')
            book = main_book.active
            break
        except Exception:
            print('Uh Oh! Error occurred, try again.')
            pass


    def word_finder(searchString):
        now = datetime.datetime.now()
        current_time = now.strftime("%H:%M")
        for i in range(1, book.max_row + 1):
            for j in range(1, book.max_column + 1):
                if searchString == book.cell(i, j).value:
                    book.cell(i, 4).value = current_time
                    main_book.save(f'{path}/{topic}_attendance.xlsx')
                else:
                    pass


    @app.route('/participant_name', methods=['POST', 'GET'])
    def append_participant():
        global current_line
        try:
            if flask.request.method == 'GET':
                name = flask.request.args.get('name')
                roll_no = flask.request.args.get('roll_no')
                print(f"\n-> {name}, roll no. {roll_no} has joined the class.")
                now = datetime.datetime.now()
                current_time = now.strftime("%H:%M")
                participants.append([name, str(current_time)])
                book[f'A{current_line}'] = name
                book[f'B{current_line}'] = roll_no
                book[f'C{current_line}'] = current_time
                current_line += 1
                main_book.save(f'{path}/{topic}_attendance.xlsx')
            return f'Hi {name} Your name has been noted succesfully'
        except Exception:
            return 'Invalid request'


    @app.route('/left_meeting', methods=['POST', 'GET'])
    def note_end_meeting_time():
        name = flask.request.args.get('name_end')
        word_finder(name)
        print(f"\n-> {name} has left the class.")
        return 'Success'


    @app.route('/not_attentive', methods=['POST', 'GET'])
    def notify_note_attentive():
        name = flask.request.args.get('name')
        roll_no = flask.request.args.get('roll_no')
        notification.notify(
            title=f'{name} , roll number {roll_no} is Not listening',
            message=f'Call him/her, Do it now!',
            app_icon=r'resources/logo.ico',
            timeout=5

        )
        return f"notified"


    ngrok.set_auth_token('1oKXCEyBJBZzLjaxlScfawQsFBH_HeUsFkWjrcUMUYWpzfHS')
    url = ngrok.connect(8085).public_url
    print(f'Code: {url.strip("http://").strip(".ngrok.io")}')
    app.run('0.0.0.0', 8085)

elif mode == 'J' or mode == 'j':
    code = input('Enter the code given by teacher: ')
    while 1:
        enter_name = input('Type your name: ')
        enter_roll_no = input('Type your Roll Number: ')
        if enter_roll_no == '' or enter_roll_no == ' ':
            print("Hey enter your name!!!")
            pass
        else:
            break
    req = requests.get('http://' + code + '.ngrok.io' + f'/participant_name?name={enter_name}&roll_no={enter_roll_no}')
    face_check = 0
    bool_one = False
    frame = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    frame.set(10, 150)
    faceCascade = cv2.CascadeClassifier('resources/haarcascade.xml')
    print('Press "q" to exit')
    while True:
        check, vid = frame.read()
        imgGray = cv2.cvtColor(vid, cv2.COLOR_BGR2GRAY)
        faces = faceCascade.detectMultiScale(imgGray, 1.3, 5)
        if len(faces) > 0:
            face_check = 0
            bool_one = False

        else:
            face_check = face_check + 1
        if face_check >= 150:
            if bool_one:
                continue
            req_nt = requests.get(
                'http://' + code + '.ngrok.io' + f'/not_attentive?name={enter_name}&roll_no={enter_roll_no}')

            print('check')
            bool_one = True
        for (x, y, w, h) in faces:
            img = cv2.rectangle(vid, (x, y), (x + w, y + h), (255, 0, 0), 2)
        cv2.imshow('Your video', vid)
        # the 'q' button is set as the
        # quitting button you may use any
        # desired button of your choice
        if cv2.waitKey(1) & 0xFF == ord('q'):
            req = requests.get(
                'http://' + code + '.ngrok.io' + f'/left_meeting?name_end={enter_name}')
            print('Your leave time is being noted...please wait')
            print('Done!')
            break

    frame.release()
    cv2.destroyAllWindows()
